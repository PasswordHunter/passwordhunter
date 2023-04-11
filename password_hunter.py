import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from edge import edge_browser
from chrome import Chrome
from firefox_plus import FirefoxBrowser
from brave import Brave
from windpass import WindowWifiPass
import subprocess,os
from openpyxl import Workbook

winpass = WindowWifiPass()
# firefox data
firefox1 = FirefoxBrowser()
# brave data
brave = Brave()
# chrome info variables
chromes = Chrome()
password_data = chromes.passwords()
cookies = chromes.cookies()
web_history = chromes.history()
search_terms= chromes.search_terms()
autofill_data = chromes.web_data()
credit_card_info= chromes.credit_card_chrome()
downloads=chromes.get_chrome_downloads()
# main pages

class Main:
    def __init__(self, master):
        self.master = master
        self.path_icon = "icon.ico"
        master.title("Password Hunter")
        master.iconbitmap(self.path_icon)

        # create label widget
        self.label = tk.Label(master, text="Password Hunter", font=("Helvetica", 14,"bold"))
        self.label.pack()

        # create listbox widget
        self.listbox = tk.Listbox(master, font=("Helvetica", 12,"bold"), height=15, selectbackground="grey")
        self.listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)


        self.password_label = tk.Label(root, font=("Helvetica", 12,"bold"),fg="white", text="")
        self.password_label.pack_forget()

        self.button = tk.Button(root, text="Scan for Network Profiles",
                                command=self.scan_wifi_profiles, justify='left', bg='red',font=("Helvetica", 12, "bold"),fg="white")
        self.button.pack_forget()
        
        self.button2 = tk.Button(
            root, text="Show Password",justify='left', bg='green',font=("Helvetica", 12, "bold"),fg="white", command=self.profile_show_password)
        self.button2.pack_forget()
        #button for password 
        self.password_button = tk.Button(root, text="Show Password",font=("Helvetica", 12, "bold"),bg="green",fg="white", command=self.show_password)
        self.password_button.pack_forget()
        self.hide_password_button = tk.Button(root, text="Hide Passwords",font=("Helvetica", 12, "bold"),bg="red",fg="white", command=self.hide_chrome_password_)
        self.hide_password_button.pack_forget()
        
        self.edge_password_button = tk.Button(root, text="Show Password",font=("Helvetica", 12, "bold"),bg="green",fg="white", command=self.show_edge_password)
        self.edge_password_button.pack_forget()
        self.edge_hide_password_button = tk.Button(root, text="Hide Passwords",font=("Helvetica", 12, "bold"),bg="red",fg="white", command=self.edge_hide_password)
        self.edge_hide_password_button.pack_forget()
        
        self.firefox_password_button = tk.Button(root, text="Show Password",font=("Helvetica", 12, "bold"),bg="green",fg="white", command=self.show_firefox_password)
        self.firefox_password_button.pack_forget()
        self.firefox_hide_password_button = tk.Button(root, text="Hide Passwords",font=("Helvetica", 12, "bold"),bg="red",fg="white", command=self.firefox_hide_password)
        self.firefox_hide_password_button.pack_forget()
        
        self.brave_password_button = tk.Button(root, text="Show Password",font=("Helvetica", 12, "bold"),bg="green",fg="white", command=self.brave_show_password)
        self.brave_password_button.pack_forget()
        self.brave_hide_password_button = tk.Button(root, text="Hide Passwords",font=("Helvetica", 12, "bold"),bg="red",fg="white", command=self.brave_hide_password)
        self.brave_hide_password_button.pack_forget()
        
        
        #export button
        self.chrome_export_credential = tk.Button(root,text="Export to excel file", font=("Helvetica", 12, "bold"), bg="blue", fg="white", command=self.chrome_export_credentials)
        self.firefox_export_credential = tk.Button(root,text="Export to excel file", font=("Helvetica", 12, "bold"), bg="blue", fg="white", command= self.firefox_export_credentials)
        self.edge_export_credential = tk.Button(root,text="Export to excel file", font=("Helvetica", 12, "bold"), bg="blue", fg="white", command=self.edge_export_credentials)
        self.brave_export_credential = tk.Button(root,text="Export to excel file", font=("Helvetica", 12, "bold"), bg="blue", fg="white", command=self.brave_export_credentials)
    
        # create menu widget
        self.menu = tk.Menu(master)
        master.config(menu=self.menu)

        # create chrome browser submenu
        self.chrome_menu = tk.Menu(self.menu, tearoff=False)
        self.chrome_menu.add_command(
            label="Credentials", command=self.chromecredentials)
        self.chrome_menu.add_command(
            label="History", command=self.chrome_history)
        self.chrome_menu.add_command(
            label="Cookies", command=self.chrome_cookies)
        self.chrome_menu.add_command(
            label="Autofill", command=self.chrome_autofill)
        self.chrome_menu.add_command(
            label="Credits card", command=self.chrome_credit)
        self.chrome_menu.add_command(
            label="Search Terms", command=self.chrome_seach)
        self.chrome_menu.add_command(
            label="Downloads", command=self.get_chrome_download)

        # create Edge browser submenu
        self.edge_menu = tk.Menu(self.menu, tearoff=False)
        self.edge_menu.add_command(
            label="Credentials", command=self.edge_credentials)
        self.edge_menu.add_command(label="History", command=self.edge_history)
        self.edge_menu.add_command(label="Cookies", command=self.edge_cookie)
        self.edge_menu.add_command(
            label="Downlaods", command=self.edge_downloaded)

        # create Firefox browser submenu
        self.firefox_menu = tk.Menu(self.menu, tearoff=False)
        self.firefox_menu.add_command(
            label="Credentials", command=self.firefox_credentials)
        self.firefox_menu.add_command(
            label="History", command=self.firefox_hsitories)
        self.firefox_menu.add_command(
            label="Cookies", command=self.firefox_cookie)
        self.firefox_menu.add_command(
            label="Downloads", command=self.get_firefox_download)
       
        # create brave browser submenu
        self.brave_menu = tk.Menu(self.menu, tearoff=False)
        self.brave_menu.add_command(
            label="Credentials", command=self.brave_credentials)
        self.brave_menu.add_command(label="Cookies", command=self.brave_cookie)
        self.brave_menu.add_command(
            label="History", command=self.brave_history)
        self.brave_menu.add_command(
            label="Downloads", command=self.get_brave_download)
        

        # create wifi submenu
        self.wifi_menu = tk.Menu(self.menu, tearoff=False)
       
        self.wifi_menu.add_command(
            label="Credentials", command=self.disconnect)

        self.main_menu = tk.Menu(self.menu, tearoff=False)
    

        # add submenus to main menu
        self.menu.add_cascade(label="Chrome", menu=self.chrome_menu, font=("Helvetica", 14))
        self.menu.add_cascade(label="Edge", menu=self.edge_menu)
        self.menu.add_cascade(label="FireFox", menu=self.firefox_menu)
        self.menu.add_cascade(label="Brave", menu=self.brave_menu)
        self.menu.add_cascade(label="Wifi", menu=self.wifi_menu)
        self.menu.add_cascade(label="Home", command=self.home)
        
        try:
            self.edge_password_lines = [line for line in  edge_browser.get_edge_creds().split('\n') if not line.startswith("Password")]
            self.password_lines_hide = [line for line in  password_data.split('\n') if not line.startswith("Password")]
            self.firefox_password_lines = [line for line in  firefox1.get_password().split("\n") if not line.startswith("Password")]
            self.brave_password_lines = [line for line in  brave.brave_passwords().split("\n") if not line.startswith("Password")]
        except Exception as e:
            messagebox.showerror("Error", e)
        
        #call home function
        self.home()
#installed browser
    def get_installed_browsers(self):
        try:
            default_dir_paths = {
                "Chrome": "C:\\Program Files (x86)\\Google\\Chrome\\Application\\chrome.exe",
                "Firefox": "C:\\Program Files\\Mozilla Firefox\\firefox.exe",
                "Microsoft Edge": "C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe",
                "Opera": "C:\\Program Files\\Opera\\launcher.exe",
                "Safari": "C:\\Program Files\\Safari\\Safari.exe",
                "Brave": "C:\\Program Files\\BraveSoftware\\Brave-Browser\\Application\\brave.exe",
            }
            installed_browsers = " "
            for browser_name, browser_path in default_dir_paths.items():
                if os.path.exists(browser_path):
                    if installed_browsers:
                        installed_browsers += " --- "
                        installed_browsers += browser_name
            return installed_browsers
        except Exception as e:
            messagebox.showerror("Error", e)
            
    def home(self):
        self.unpack_data()
        self.listbox.delete(0, tk.END)
        message = '''----------------------------------------------------------------------------------------------------------------
        P A S S W O R D  H U N T E R
        Version: 1.0

        Developers:

        1. KWITEE D. GAYLAH
        2. ZONGO WEND-BENEDO SIMEON

        <Source Code> https://github.com/PasswordHunter/passwordhunter

        Licensed under MIT LICENSE
        
        Please be advised that the Password Hunter program is strictly intended for 
        educational purposes only. The tool is designed to extract sensitive information 
        from popular web browsers and devices and is aimed at cyber security students and 
        professionals interested in web browser and device security.

        It is important to note that using this tool on any system or device without the 
        explicit permission of the owner is illegal and considered unethical. 
        The creators of this tool are not responsible for any unlawful or unethical 
        carried out using this program.

        We strongly encourage you to use this tool in a controlled and responsible manner, with 
        the utmost respect for the privacy and security of others. Any misuse of this tool can 
        lead to serious consequences and legal action against you
        
        Browsers installed into the computer are:
        
        '''+ self.get_installed_browsers() + ''' 
        
        ---------------------------BEST OF LUCK------------------------------------'''
        self.listbox.insert(tk.END, *message.split("\n"))
        
# unpack data 
    def unpack_data(self):
        try:
            self.button2.pack_forget()
            self.button.pack_forget()
            self.password_label.pack_forget()
            self.password_label.config(text="")
            self.password_button.pack_forget()
            self.hide_password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.edge_hide_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.edge_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
            
        except Exception as e:
            messagebox.showerror("Error",e)
            
# packs data 
    def pack_data(self):
        try:
            self.password_label.pack(pady=10)
            self.button2.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
        except Exception as e:
            messagebox.showerror("Error",e)
            
# shows chrome password          
    def show_password(self):
        self.listbox.delete(0, tk.END)
        try:
            for line in password_data.split("\n"):
                self.listbox.insert(tk.END, line)
                if line.startswith("Password"):
                    self.listbox.itemconfig(tk.END, fg="white", bg="green")
            # self.listbox.insert(tk.END, *self.password_lines_show)
            self.chrome_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.hide_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.password_button.pack_forget()
            self.button2.pack_forget()
            self.button.pack_forget()
            self.password_label.pack_forget()
            self.password_label.config(text="")
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
        except Exception as e:
            messagebox.showerror('Error', e)
           
#chrome export credentials
    def chrome_export_credentials(self):
        try:
            if chromes.get_installed_browsers():
                filepath = filedialog.asksaveasfilename(defaultextension='.xlsx')
                # Return if the user cancels the dialog box
                if not filepath:
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(['URL', 'Username', 'Password', 'Creation Date', 'Last Used'])
                password_data = chromes.chrome_passwords_to_excel()
                for row in password_data:
                    ws.append(row)
                wb.save(filepath)
                messagebox.showinfo('Saved', 'Your credentials data have been saved to '+filepath)
            else:
                messagebox.showwarning('Not Found', "Chrome browser is not installed!") 
        except Exception as e:
            messagebox.showerror('Error', e) 

#brave export credentials
    def brave_export_credentials(self):
        try:
            if brave.get_installed_browsers():
                filepath = filedialog.asksaveasfilename(defaultextension='.xlsx')
                # Return if the user cancels the dialog box
                if not filepath:
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(['URL', 'Username', 'Password',"Creation date","Last used","Last modified"])
                password_data = brave.brave_passwords_to_excel()
                for row in password_data:
                    ws.append(row)
                wb.save(filepath)
                messagebox.showinfo('Saved', 'Your credentials data have been saved to '+filepath)
            else:
                messagebox.showwarning('Not Found', "Brave browser is not installed!") 
        except Exception as e:
            messagebox.showerror('Error', e)            

#edge export credentials
    def edge_export_credentials(self):
        try:
            if edge_browser.get_edge_install():
                filepath = filedialog.asksaveasfilename(defaultextension='.xlsx')
                # Return if the user cancels the dialog box
                if not filepath:
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(['Index','URL', 'Username', 'Password',"Creation date"])
                password_data = edge_browser.edge_credentials_to_excel()
                for row in password_data:
                    ws.append(row)
                wb.save(filepath)
                messagebox.showinfo('Saved', 'Your credentials data have been saved to '+filepath)
            else:
                messagebox.showwarning('Not Found', "Edge browser is not installed!") 
        except Exception as e:
            messagebox.showerror('Error', e) 
            
    def firefox_export_credentials(self):
        try:
            if firefox1.get_installed_browsers():
                filepath = filedialog.asksaveasfilename(defaultextension='.xlsx')
                # Return if the user cancels the dialog box
                if not filepath:
                    return
                wb = Workbook()
                ws = wb.active
                ws.append(['URL', 'Username', 'Password'])
                password_data = firefox1.firefox_password_to_excel()
                for row in password_data:
                    ws.append(row)
                wb.save(filepath)
                messagebox.showinfo('Saved', 'Your credentials data have been saved to '+filepath)
            else:
                messagebox.showwarning('Not Found', "Edge browser is not installed!") 
        except Exception as e:
            messagebox.showerror('Error', e) 

# hide chrome password        
    def hide_chrome_password_(self):
        try:
            self.chromecredentials()
            self.hide_password_button.pack_forget()
            
        except Exception as e:
            messagebox.showerror('Error', e)
        
#Edge show password 
    def show_edge_password(self):
        self.listbox.delete(0, tk.END)
        try:
            for line in edge_browser.get_edge_creds().split("\n"):
                self.listbox.insert(tk.END, line)
                if line.startswith("Password"):
                    self.listbox.itemconfig(tk.END, fg="white", bg="green")
            # self.listbox.insert(tk.END, *self.password_lines_show)
            self.edge_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.edge_hide_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.button2.pack_forget()
            self.button.pack_forget()
            self.password_label.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
            self.password_label.config(text="")
        except Exception as e:
            messagebox.showerror('Error', e)
      
#Edge hide password      
    def edge_hide_password(self):
        try:
            self.edge_credentials()   
            self.edge_hide_password_button.pack_forget()
        except Exception as e:
            messagebox.showerror('Error', e)

#Firefox show password            
    def show_firefox_password(self):
        self.listbox.delete(0, tk.END)
        try:
            for line in firefox1.get_password().split("\n"):
                self.listbox.insert(tk.END, line)
                if line.startswith("Password"):
                    self.listbox.itemconfig(tk.END, fg="white", bg="green")
            # self.listbox.insert(tk.END, *self.password_lines_show)
            self.firefox_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.firefox_hide_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.button2.pack_forget()
            self.button.pack_forget()
            self.password_label.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.edge_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.password_label.config(text="")
        except Exception as e:
            messagebox.showerror('Error', e) 

#Firefox hide password           
    def firefox_hide_password(self):
        try:
            self.firefox_credentials()   
            self.firefox_hide_password_button.pack_forget() 
        except Exception as e:
            messagebox.showerror('Error', e)

#Brave show password      
    def brave_show_password(self):
        self.listbox.delete(0, tk.END)
        try:
            for line in brave.brave_passwords().split("\n"):
                self.listbox.insert(tk.END, line)
                if line.startswith("Password"):
                    self.listbox.itemconfig(tk.END, fg="white", bg="green")
            # self.listbox.insert(tk.END, *self.password_lines_show)
            self.brave_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.brave_hide_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            self.password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.edge_export_credential.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.button2.pack_forget()
            self.button.pack_forget()
            self.password_label.pack_forget()
            self.password_label.config(text="")
        except Exception as e:
            messagebox.showerror('Error', e)

#Brave hide password          
    def brave_hide_password(self):
        try:
            self.brave_credentials()
            self.brave_hide_password_button.pack_forget() 
        except Exception as e:
            messagebox.showerror('Error', e)

#Chrome credentials function         
    def chromecredentials(self):
        self.label.config(text="Chrome Credentials storages")
        try: 
            if chromes.get_installed_browsers():
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *self.password_lines_hide)
                self.chrome_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                self.password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                self.button2.pack_forget()
                self.button.pack_forget()
                self.password_label.pack_forget()
                self.edge_password_button.pack_forget()
                self.edge_hide_password_button.pack_forget()
                self.password_label.config(text="")
                self.firefox_password_button.pack_forget()
                self.firefox_hide_password_button.pack_forget()
                self.brave_password_button.pack_forget()
                self.brave_hide_password_button.pack_forget()
                self.firefox_export_credential.pack_forget()
                self.edge_export_credential.pack_forget()
                self.brave_export_credential.pack_forget()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)

# chrome cookies functions
    def chrome_cookies(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome Cookies selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *cookies.splitlines())
                self.unpack_data()
                self.password_button.pack_forget()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)

# Chrome downloads
    def get_chrome_download(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome Downloads selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *downloads.splitlines())
                self.unpack_data()
                self.password_button.pack_forget()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)
#chrome history functions
    def chrome_history(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome History selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *web_history.splitlines())
                self.unpack_data()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)
            
#chrome search term funtion
    def chrome_seach(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome Search terms selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *search_terms.splitlines())
                self.unpack_data()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)
            
# chrome autofill function
    def chrome_autofill(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome autofill information")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *autofill_data.splitlines())
                self.unpack_data()
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror('Error', e)

# chrome credit card function
    def chrome_credit(self):
        try:
            if chromes.get_installed_browsers():
                self.label.config(text="Chrome credit cart information")
                self.unpack_data()
                if len(credit_card_info) != 0:
                    self.listbox.delete(0, tk.END)
                    self.listbox.insert(tk.END, *credit_card_info)
                else:
                    self.listbox.delete(0, tk.END)
                    messagebox.showinfo(
                        title="Info", message="No credit card information available")
            else:
                messagebox.showinfo("Info", "Chrome Browser is not installed")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")

#firefox credentials functions 
    def firefox_credentials(self):
        try:
            if firefox1.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Firefox Credentials selected")
                self.button2.pack_forget()
                self.button.pack_forget()
                self.password_label.pack_forget()
                self.edge_password_button.pack_forget()
                self.edge_hide_password_button.pack_forget()
                self.password_label.config(text="")
                self.brave_password_button.pack_forget()
                self.brave_hide_password_button.pack_forget()
                self.chrome_export_credential.pack_forget()
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *self.firefox_password_lines)
                self.firefox_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                self.firefox_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            else:
                messagebox.showinfo("Info", "Firefox Browser is not installed")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")

# firefox downloads
    def get_firefox_download(self):
        try:
            if firefox1.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Firefox Downloads selected")
                self.listbox.delete(0, tk.END)
        
                self.listbox.insert(
                    tk.END, *firefox1.get_firefox_downloads().splitlines())
            else:
                messagebox.showinfo("Info", "Firefox Browser is not installed")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
#firefox cookies functions
    def firefox_cookie(self):
        try:
            if firefox1.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Firefox Cookies selected")
                self.listbox.delete(0, tk.END)
        
                self.listbox.insert(
                    tk.END, *firefox1.firefox_cookies().splitlines())
            else:
                messagebox.showinfo("Info", "Firefox Browser is not installed")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            
#firefox histores function
    def firefox_hsitories(self):
        try:
            if firefox1.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Firefox History selected")
                self.listbox.delete(0, tk.END)
            
                self.listbox.insert(
                    tk.END, *firefox1.firefox_history().splitlines())
            else:
                messagebox.showinfo("Info", "Firefox Browser is not installed")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            
# edge credentials functions
    def edge_credentials(self):
        try:
            if edge_browser.get_edge_install():
                self.unpack_data()
                self.label.config(text="Microsoft Edge credentials selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(
                    tk.END, *self.edge_password_lines)
                self.edge_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                self.edge_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                self.button2.pack_forget()
                self.button.pack_forget()
                self.password_label.pack_forget()
                self.firefox_password_button.pack_forget()
                self.firefox_hide_password_button.pack_forget()
                self.brave_password_button.pack_forget()
                self.brave_hide_password_button.pack_forget()
                self.password_label.config(text="")
            else:
                messagebox.showinfo("Info", "Microsoft Edge Browser is not installed")
        except Exception as e:
            self.listbox.delete(0, tk.END)
            messagebox.showerror("Error", f"[E] {str(e)}")
            
#edge cookies functions
    def edge_cookie(self):
        try:
            if edge_browser.get_edge_install():
                self.label.config(text="Microsoft Edge Cookies selected")
                self.unpack_data()
            
                self.listbox.delete(0, tk.END)
                self.listbox.insert(
                    tk.END, *edge_browser.edge_cookies().splitlines())
            else:
                messagebox.showinfo("Info", "Microsoft Edge Browser is not installed")
        except Exception as e:
            self.listbox.delete(0, tk.END)
            messagebox.showerror("Error", f"[E] {str(e)}")
            
#edge history
    def edge_history(self):
        try:
            if edge_browser.get_edge_install():
                self.unpack_data()
                self.label.config(text="Microsoft Edge History selected")
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *edge_browser.edge_history())
            else:
                messagebox.showinfo("Info", "Microsoft Edge Browser is not installed")
        except Exception as e:
            self.listbox.delete(0, tk.END)
            messagebox.showerror("Error", f"[E] {str(e)}")
            
#edge download
    def edge_downloaded(self):
        try:
            if edge_browser.get_edge_install():
                self.unpack_data()
                self.label.config(text=" Microsoft Edge downloads selected")
            
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *edge_browser.edge_downloads())
            else:
                messagebox.showinfo("Info", "Microsoft Edge Browser is not installed")
        except Exception as e:
            self.listbox.delete(0, tk.END)
            messagebox.showerror("Error", f"[E] {str(e)}")

# brave credentials functionality
    def brave_credentials(self):
        try:
            if brave.get_installed_browsers():
                self.unpack_data()
                
                self.label.config(text="Brave credentials selected")
                self.listbox.delete(0, tk.END)
                if brave.kill_program("brave.exe"):
                    messagebox.showinfo(
                        "Running", "Brave browser was running. It has been killed")
                    self.listbox.insert(
                        tk.END, *self.brave_password_lines)
                    self.brave_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                    self.brave_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                else:
                    self.listbox.insert(
                        tk.END, *self.brave_password_lines)
                    self.brave_export_credential.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
                    self.brave_password_button.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            else:
                messagebox.showwarning(
                    "Warning", "Brave browser is not installed!")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            
# brave history function 
    def brave_history(self):
        try:
            if brave.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Brave History selected")
      
                if brave.kill_program("brave.exe"):
                    messagebox.showinfo(
                        "Running", "Brave browser was running. It has been killed")
                    self.listbox.insert(
                        tk.END, *brave.get_brave_history().splitlines())
                else:
                    self.listbox.delete(0, tk.END)
                    self.listbox.insert(
                        tk.END, *brave.get_brave_history().splitlines())
            else:
                self.listbox.delete(0, tk.END)
                messagebox.showinfo(
                    "Info", "Brave browser is not installed!")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            self.listbox.delete(0, tk.END)
            
# brave cookies function
    def brave_cookie(self):
        try:
            if brave.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Brave Cookies selected")
                if brave.kill_program("brave.exe"):
                    messagebox.showinfo(
                        "Running", "Brave browser was running. It has been killed")
                    self.listbox.insert(
                        tk.END, *brave.brave_cookies().splitlines())
                else:
                    self.listbox.delete(0, tk.END)
                    self.listbox.insert(
                        tk.END, *brave.brave_cookies().splitlines())
            else:
                self.listbox.delete(0, tk.END)
                messagebox.showinfo(
                    "Info", "Brave browser is not installed!")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            self.listbox.delete(0, tk.END)
            
# brave downloads function
    def get_brave_download(self):
        try:
            if brave.get_installed_browsers():
                self.unpack_data()
                self.label.config(text="Brave Downloads selected")
                if brave.kill_program("brave.exe"):
                    messagebox.showinfo(
                        "Running", "Brave browser was running. It has been killed")
                    self.listbox.insert(
                        tk.END, *brave.get_brave_downloads().splitlines())
                else:
                    self.listbox.delete(0, tk.END)
                    self.listbox.insert(
                        tk.END, *brave.get_brave_downloads().splitlines())
            else:
                self.listbox.delete(0, tk.END)
                messagebox.showinfo(
                    "Info", "Brave browser is not installed!")
        except Exception as e:
            messagebox.showerror("Error", f"[E] {str(e)}")
            self.listbox.delete(0, tk.END)

# Wifi ---- info
    def disconnect(self):
        try:
            self.label.config(text="Wifi Profiles")
            self.listbox.delete(0, tk.END)
            self.listbox.insert(tk.END, *winpass.get_devices_info().splitlines())
            self.button.pack(side=tk.LEFT, pady=20, anchor=tk.CENTER, padx=10)
            self.password_button.pack_forget()
            self.hide_password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.edge_hide_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.edge_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
        except Exception as e:
            messagebox.showerror('Error',str)

# Wifi ---- scan profiles function
    def scan_wifi_profiles(self):
        try:
            self.password_button.pack_forget()
            self.hide_password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.edge_hide_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.edge_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
            self.password_label.pack(pady=10)
            self.button2.pack(side=tk.RIGHT, pady=20, anchor=tk.CENTER, padx=10)
            try:
                self.listbox.delete(0, tk.END)
                self.listbox.insert(tk.END, *winpass.get_wifi_profile())
            except Exception as e:
                messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error", str(e))

# Wifi ---- show password function          
    def profile_show_password(self):
        try: 
            self.password_button.pack_forget()
            self.hide_password_button.pack_forget()
            self.edge_password_button.pack_forget()
            self.edge_hide_password_button.pack_forget()
            self.firefox_password_button.pack_forget()
            self.firefox_hide_password_button.pack_forget()
            self.brave_password_button.pack_forget()
            self.brave_hide_password_button.pack_forget()
            self.chrome_export_credential.pack_forget()
            self.firefox_export_credential.pack_forget()
            self.edge_export_credential.pack_forget()
            self.brave_export_credential.pack_forget()
            selected_item = self.listbox.get(self.listbox.curselection())
            try:
                if selected_item is not None:
                    output = subprocess.check_output(
                        ["netsh", "wlan", "show", "profile", selected_item, "key=clear"], shell=True)
                    output = output.decode("utf-8")
                    password_line = [line.strip() for line in output.split(
                        "\n") if "Key Content" in line][0]
                    password = password_line.split(":")[1].strip()
                    self.password_label.config(
                        text=f"Password for: {selected_item} ==> {password}", background="black")
                    tk.Button(self.password_label, text="Copy",
                            command=lambda: root.clipboard_append(password))
                else:   
                    messagebox.showwarning("Profile not selected", "Please select Wi-Fi profile!!!")
            except Exception as e:
                messagebox.showerror(title=selected_item, message=str(e))
        except Exception as e:
            messagebox.showwarning("Profile not selected", "Please select Wi-Fi profile!!!")

# Window display function 
def center_window(root, window_width, window_height):
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = int((screen_width/2) - (window_width/2))
    y = int((screen_height/2) - (window_height/2))
    root.geometry(f'{window_width}x{window_height}+{x}+{y}')


if __name__ == "__main__":
    root = tk.Tk()
    menu = Main(root)
    center_window(root, 800, 500)
    root.minsize(800, 500)
    root.mainloop()
